import requests
import json
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ── 설정 ────────────────────────────────────────────────────
SERVICE_KEY = "0104IX0P98831DDJ"
BASE_URL    = "https://www.safetydata.go.kr/V2/api/DSSP-IF-00247"
MAX_ROWS    = 1000

MASTER_XLSX = "재난문자_전체.xlsx"
STATE_FILE  = "collection_state.json"
CSV_SOURCE  = r"c:\Users\namzx\Downloads\행정안전부_긴급재난문자.csv"

# (API 필드명, 한국어 헤더, 열 너비)
COLUMNS = [
    ("SN",           "일련번호",   12),
    ("CRT_DT",       "생성일시",   20),
    ("MSG_CN",       "메시지내용", 60),
    ("RCPTN_RGN_NM", "수신지역명", 40),
    ("EMRG_STEP_NM", "긴급단계명", 12),
    ("DST_SE_NM",    "재해구분명", 12),
    ("REG_YMD",      "등록일시",   14),
    ("MDFCN_YMD",    "수정일시",   14),
]

API_KEYS  = [c[0] for c in COLUMNS]
KOR_NAMES = [c[1] for c in COLUMNS]


# ── 상태 파일 ────────────────────────────────────────────────

def load_state():
    p = Path(STATE_FILE)
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return None


def save_state(max_sn, api_total_count, excel_total_rows):
    state = {
        "max_sn":           max_sn,
        "api_total_count":  api_total_count,
        "excel_total_rows": excel_total_rows,
        "last_run":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    print(f"상태 저장 완료 → max_sn={max_sn:,}, 총 {excel_total_rows:,}행")


# ── API 수집 ─────────────────────────────────────────────────

def fetch_page(page_no):
    params = {
        "serviceKey": SERVICE_KEY,
        "pageNo":     page_no,
        "numOfRows":  MAX_ROWS,
        "returnType": "json",
    }
    resp = requests.get(BASE_URL, params=params, timeout=15)
    resp.raise_for_status()
    return resp.json()


def parse_response(data):
    header = data.get("header", {})
    if header.get("resultCode") != "00":
        raise ValueError(f"API 오류: {header.get('resultMsg', '알 수 없는 오류')}")
    total_count = int(data.get("totalCount", 0))
    rows = data.get("body") or []
    return total_count, rows


def fetch_pages(start_page=1):
    """start_page부터 마지막 페이지까지 수집."""
    collected = []
    page_no   = start_page
    api_total = 0

    while True:
        try:
            data = fetch_page(page_no)
            api_total, rows = parse_response(data)
        except requests.RequestException as e:
            print(f"  [네트워크 오류] 페이지 {page_no}: {e}")
            break
        except (ValueError, KeyError, TypeError) as e:
            print(f"  [파싱 오류] 페이지 {page_no}: {e}")
            break

        if not rows:
            break

        collected.extend(rows)
        print(f"  페이지 {page_no:>3} | {len(rows):>4}건 | "
              f"이번 수집 누적 {len(collected):>6} | API 전체 {api_total:,}건")

        if page_no * MAX_ROWS >= api_total:
            break

        page_no += 1
        time.sleep(0.3)

    return api_total, collected


# ── CSV 로드 ─────────────────────────────────────────────────

def load_csv_source():
    """기존 CSV 파일을 API와 동일한 dict 형태로 로드."""
    p = Path(CSV_SOURCE)
    if not p.exists():
        print(f"CSV 없음 (건너뜀): {p}")
        return []

    # 2번째 행 = 한국어 설명 행 → skiprows=[1] 로 제거
    df = pd.read_csv(p, encoding="utf-8-sig", skiprows=[1], dtype=str)
    df.columns = df.columns.str.strip()

    # 유효한 행만 (SN이 숫자)
    df = df[df["SN"].str.match(r"^\d+$", na=False)].copy()

    # API 필드명 기준으로 컬럼 정렬
    existing_cols = [k for k in API_KEYS if k in df.columns]
    df = df[existing_cols]

    items = df.to_dict("records")
    print(f"CSV 로드: {len(items):,}건 ({p.name})")
    return items


# ── 병합 ────────────────────────────────────────────────────

def merge_by_sn(base_items, extra_items):
    """SN 기준 중복 제거 후 오름차순 정렬."""
    merged = {int(item["SN"]): item for item in base_items}
    new_cnt = 0
    for item in extra_items:
        sn = int(item["SN"])
        if sn not in merged:
            merged[sn] = item
            new_cnt += 1
    sorted_items = [merged[k] for k in sorted(merged)]
    return sorted_items, new_cnt


# ── Excel ────────────────────────────────────────────────────

def _header_style(ws):
    """헤더 행(1행) 스타일 적용."""
    h_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_cf(ws, last_data_row):
    """짝수 행 색상 조건부 서식 적용 (전체 범위 재설정)."""
    from openpyxl.formatting.formatting import ConditionalFormattingList
    ws.conditional_formatting = ConditionalFormattingList()

    data_range = f"A2:{get_column_letter(len(COLUMNS))}{last_data_row}"
    even_fill  = PatternFill(fill_type="solid", fgColor="EBF3FB")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=even_fill)
    )


def write_excel(items):
    """전체 데이터로 Excel 파일을 새로 생성."""
    # 1. pandas로 데이터 빠르게 쓰기
    rows = [[str(item.get(k, "")) for k in API_KEYS] for item in items]
    df   = pd.DataFrame(rows, columns=KOR_NAMES)
    df.to_excel(MASTER_XLSX, index=False, engine="openpyxl")

    # 2. openpyxl로 스타일 적용
    wb = load_workbook(MASTER_XLSX)
    ws = wb.active
    ws.title = "재난문자"

    ws.sheet_format.defaultRowHeight = 40
    _header_style(ws)
    _apply_cf(ws, last_data_row=len(items) + 1)
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(MASTER_XLSX)
    print(f"Excel 생성 완료: {MASTER_XLSX} ({len(items):,}건)")


def append_excel(new_items):
    """기존 Excel에 새 행을 추가."""
    wb = load_workbook(MASTER_XLSX)
    ws = wb.active

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    e_fill = PatternFill(fill_type="solid", fgColor="EBF3FB")
    center = Alignment(horizontal="center", vertical="center")
    wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    start_row = ws.max_row + 1
    for i, item in enumerate(new_items):
        row_idx = start_row + i
        for col_idx, key in enumerate(API_KEYS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(item.get(key, "")))
            cell.alignment = wrap if col_idx == 3 else center
            cell.border    = border
            if row_idx % 2 == 0:
                cell.fill = e_fill

    # 조건부 서식 범위 및 자동 필터 갱신
    _apply_cf(ws, last_data_row=ws.max_row)
    ws.auto_filter.ref = ws.dimensions

    wb.save(MASTER_XLSX)
    print(f"Excel 추가 완료: {MASTER_XLSX} (+{len(new_items):,}건)")


# ── 메인 ────────────────────────────────────────────────────

def main():
    sep = "=" * 55
    print(sep)
    state = load_state()

    # ── 최초 실행 ────────────────────────────────────────────
    if state is None:
        print("[ 최초 실행 ] 전체 데이터 수집 시작")

        # 1) 기존 CSV 로드
        csv_items = load_csv_source()

        # 2) API 전체 수집 (1페이지부터)
        print("API 전체 수집 중...")
        api_total, api_items = fetch_pages(start_page=1)

        # 3) 병합 (SN 중복 제거, 오름차순 정렬)
        all_items, _ = merge_by_sn(csv_items, api_items)
        print(f"병합 결과: 총 {len(all_items):,}건 "
              f"(CSV {len(csv_items):,} + API {len(api_items):,}건, 중복 제거 포함)")

        # 4) Excel 저장
        print("Excel 파일 생성 중...")
        write_excel(all_items)

        # 5) 상태 저장
        max_sn = max(int(item["SN"]) for item in all_items)
        save_state(max_sn, api_total, len(all_items))

    # ── 이후 실행 (증분 수집) ────────────────────────────────
    else:
        prev_max_sn    = state["max_sn"]
        prev_api_total = state["api_total_count"]
        prev_rows      = state["excel_total_rows"]

        print(f"[ 증분 실행 ] 이전 수집: {prev_rows:,}건 "
              f"| max_sn={prev_max_sn:,} | 마지막 실행: {state.get('last_run', '-')}")

        # 새 레코드가 있을 가능성이 있는 페이지부터 시작 (1페이지 여유)
        start_page = max(1, prev_api_total // MAX_ROWS)
        print(f"API 수집 시작 페이지: {start_page}페이지")

        api_total, fetched = fetch_pages(start_page=start_page)

        # 이미 수집한 SN 이후 데이터만 필터
        new_items = [item for item in fetched if int(item["SN"]) > prev_max_sn]

        if not new_items:
            print("새로운 데이터 없음. 이미 최신 상태입니다.")
            print(sep)
            return

        print(f"신규 데이터: {len(new_items):,}건")

        # Excel에 추가
        append_excel(new_items)

        # 상태 업데이트
        max_sn     = max(int(item["SN"]) for item in new_items)
        total_rows = prev_rows + len(new_items)
        save_state(max_sn, api_total, total_rows)

    print(sep)
    print("완료.")


if __name__ == "__main__":
    main()
