import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = r"c:\Users\namzx\Downloads\행정안전부_긴급재난문자.csv"

COLUMN_MAP = {
    "SN": "일련번호",
    "CRT_DT": "생성일시",
    "MSG_CN": "메시지내용",
    "RCPTN_RGN_NM": "수신지역명",
    "EMRG_STEP_NM": "긴급단계명",
    "DST_SE_NM": "재해구분명",
    "REG_YMD": "등록일시",
    "MDFCN_YMD": "수정일시",
}

# 각 컬럼의 Excel 열 너비 (문자 수 기준)
COLUMN_WIDTHS = {
    "일련번호": 12,
    "생성일시": 20,
    "메시지내용": 60,
    "수신지역명": 40,
    "긴급단계명": 12,
    "재해구분명": 12,
    "등록일시": 14,
    "수정일시": 14,
}


def load_csv(path):
    # 첫 행: 영문 컬럼명, 두 번째 행: 한국어 설명 행 → skiprows=[1]로 건너뜀
    df = pd.read_csv(path, encoding="utf-8-sig", skiprows=[1], dtype=str)
    df = df.rename(columns=COLUMN_MAP)

    # 실제 데이터 행만 남김 (일련번호가 숫자인 행)
    df = df[df["일련번호"].str.match(r"^\d+$", na=False)].copy()
    df = df.reset_index(drop=True)

    print(f"총 {len(df)}건 로드 완료")
    return df


def apply_excel_style(ws, df):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더 스타일
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 데이터 행 스타일
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # 메시지내용 컬럼(3번째)은 줄바꿈 허용
            cell.alignment = wrap if col_idx == 3 else center
            cell.border = border
            # 짝수 행 연한 배경
            if row_idx % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="EBF3FB")

    # 열 너비 설정
    for col_idx, col_name in enumerate(df.columns, start=1):
        width = COLUMN_WIDTHS.get(col_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 행 높이: 헤더는 22, 데이터는 45 (메시지 내용이 길어서)
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 45

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 자동 필터
    ws.auto_filter.ref = ws.dimensions


def save_to_excel(df, output_path):
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "재난문자"

    apply_excel_style(ws, df)

    wb.save(output_path)
    print(f"Excel 저장 완료: {output_path}")
    print(f"  - 행 수: {len(df)}건")
    print(f"  - 열 수: {len(df.columns)}개")


if __name__ == "__main__":
    csv_path = Path(CSV_PATH)
    if not csv_path.exists():
        print(f"파일을 찾을 수 없습니다: {csv_path}")
        raise SystemExit(1)

    output_path = csv_path.parent / (csv_path.stem + "_정리.xlsx")

    df = load_csv(csv_path)
    print(df[["일련번호", "생성일시", "재해구분명", "수신지역명"]].head(3).to_string(index=False))

    save_to_excel(df, output_path)
